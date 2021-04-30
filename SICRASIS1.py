import os
import sys
import time
import getpass
from tkinter import *
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
global nlogin2
nlogin2 = 0

def verificaloginadm():
    logins = 'logins.txt'
    senhas = 'senhas.txt'
    acesso = -1
    if os.path.isfile(logins) and os.path.isfile(senhas) :
        login2(nlogin2)
    else:
        cadastroinicialdeadmin()
        SICRA()
    
def verificaloginuser():
    logins = 'usuario.txt'
    senhas = 'passwords.txt'
    if os.path.isfile(logins) and os.path.isfile(senhas) :
        Login()
        
        
    else:
        cadastroinicialdeusuario()
        acessou =0
        SICRA()
        
def destruirjanela():
    
    rootA.destroy()

def menuprincipal():
        clear()
        print('-----------------------------------------------------------')
        print('                    MENU PRINCIPAL                         ')
        print('                      S.I.C.R.A                         \n ')

        print('SELECIONE O TIPO DE USUÁRIO:')
        print('1 - ADMINISTRADOR')
        print('2 - LOCADOR')
        print('0 - SAIR\n')

        print('-----------------------------------------------------------')
        opcao = 2
        try:

            opcao = int(input())
            print(opcao)
        except:
            clear()
            print('Opção inválida, favor digitar uma opção válida.')
            time.sleep(1)
            clear()
            return -1

        return (opcao)

def cadastroinicialdeadmin():
        acesso = -1
        clear()
        login = input('DIGITE O SEU LOGIN: ')
        senha = input('DIGITE O SUA SENHA: ')
        if login == '':
            login = '0'
        if senha == '':
            senha = '0'
        username1 = open('logins.txt','w')
        username1.write(login)
        username1.close()
        username1 = open('logins.txt','r')
        for linha in username1:
            linha = linha.rstrip()
            print(linha)
        username1.close()

        passwordarq = open('senhas.txt', 'w')
        passwordarq.write(senha)
        passwordarq.close()

        passwordarq = open('senhas.txt', 'r')
        for linha in passwordarq:
            linha = linha.rstrip()
            print(linha)
        passwordarq.close()
        
def sair():
    sys.exit()
    # quit()

def clear():
    os.system('cls' if os.name == 'nt' else clear)

def Login():
    global loginEL
    global senhaEL 
    global rootA
   
 
    rootA = Tk() 
    rootA.title('Login') 
 
    intruction = Label(rootA, text='Please Login\n') 
    intruction.grid(sticky=E) 
 
    login = Label(rootA, text='Locador: ') 
    senha = Label(rootA, text='Senha: ') 
    login.grid(row=1, sticky=W)
    senha.grid(row=2, sticky=W)
 
    loginEL = Entry(rootA) 
    senhaEL = Entry(rootA, show='*')
    loginEL.grid(row=1, column=1)
    senhaEL.grid(row=2, column=1)
 
    loginB = Button(rootA, text='Login', command=login1)
    loginB.grid(columnspan=2, sticky=W)

    # rmuser = Button(rootA, text='Fechar aba', fg='red', command=destruirjanela) 
    # rmuser.grid(columnspan=2, sticky=W)

    rootA.mainloop()

def login1():
    global acessou
    global login
    

    
    
    login =loginEL.get()
    senha = senhaEL.get()
    
    
    if login == '' or senha == '':
        clear()
        print('Login inválido')
        time.sleep(2)
        SICRA()
    else:
        acessou = 0
        #lendo o arquivo e verificando se o login e senha estão contidos nele
        logindeuser = open('usuario.txt', 'r')
        #transformando o que tem no arquivo em uma lista
        listadelogin = logindeuser.readlines()
        # print(listadelogin)
        for i in range(len(listadelogin)):
            # print(listadelogin[i])
            listadelogin[i] = listadelogin[i].split('\n')
        
        # time.sleep(2)

        senhadeuser = open('passwords.txt', 'r')
        listadesenha = senhadeuser.readlines()
        
        print(len(listadesenha))
        for i in range(len(listadesenha)):
            # print(listadesenha[i])
            listadesenha[i] = listadesenha[i].split('\n')
        # print(listadelogin)
        
        # time.sleep(2)
        # print(len(listadelogin))
        # print(len(listadesenha))
        # time.sleep(5)
        #verificando se o login e senha estão iguais aos que estão nos arquivos de login e senha
        
        for i in range(len(listadelogin)):
            for j in range(2):
                if login == listadelogin[i][j]:
                    if senha == listadesenha[i][j]:
                        clear()
                        print('----------ACESSO CONCEDIDO----------')
                        time.sleep(1)
                        acessou = 1
                        criarlog(login)
                        destruirjanela()

        if acessou == 0:
            clear()
            #senão chamando a propria função de novo para digitar a senha novamente
            print('senha incorreta por favor digite o login novamente')
            time.sleep(2)
        
    return ()

def criarlog(login) :
    data_e_hora_atuais = datetime.now()
    data_e_hora_em_texto = data_e_hora_atuais.strftime('%d/%m/%Y %H:%M')
    x = ('\n\n---------------REGISTRO-DE-LOG---------------\n')
    arquivolog = open('registrodelog.txt','a')

    arquivolog.write(x)
    arquivolog.write('O USUÁRIO {} FEZ LOGIN EM {}'.format(login,data_e_hora_em_texto))
    
    
    arquivolog.close()

def criarlogout(login):
    data_e_hora_atuais = datetime.now()
    data_e_hora_em_texto = data_e_hora_atuais.strftime('%d/%m/%Y %H:%M')
    
    arquivolog = open('registrodelog.txt','a')
    arquivolog.write('\n')
    arquivolog.write('O USUÁRIO {} FEZ LOGOUT EM {}'.format(login,data_e_hora_em_texto))
    arquivolog.write('\n---------------------------------------------\n')
    
    
    arquivolog.close()

def login2(nlogin2):
    global acesso
    login = input('Digite seu Login de administrador : ')
    senha = getpass.getpass('Digite sua senha de administrador : ')
    if login == '':
        login = '0'
    if senha == '':
        login = '0'
    
    #lendo o arquivo e verificando se o login e senha estão contidos nele
    logindeadmin = open('logins.txt', 'r')
    #transformando o que tem no arquivo em uma lista
    listadelogin = logindeadmin.readlines()
    # print(listadelogin)

    senhadeadmin = open('senhas.txt', 'r')
    listadesenha = senhadeadmin.readlines()
    
    # print(listadesenha)

    #verificando se o login e senha estão iguais aos que estão nos arquivos de login e senha
    if login in listadelogin and senha in listadesenha:
        clear()
        print('----------ACESSO CONCEDIDO----------')
        time.sleep(1)
        acesso = 1
    else:
        clear()
        #senão chamando a propria função de novo para digitar a senha novamente
        print('senha incorreta por favor digite o login novamente')
        nlogin2 = nlogin2 + 1
        if nlogin2 < 8:
            
            login2(nlogin2)
            
        else:
            clear()
            print('Limite de tentativas excedido!, tente novamente')
            time.sleep(2)
            SICRA()
    return (nlogin2)

def menudelocador(login):
            clear()
            print('-----------------------------------------------------------')
            print('MENU DO LOCADOR\n')

            print('1 - LOCAR EQUIPAMENTOS')
            print('2 - LOCAR ARMAMENTOS')
            print('3 - VER INVENTÁRIO DISPONÍVEL')
            print('4 - VOLTAR')
            print('0 - SAIR\n')

            print('-----------------------------------------------------------')
            try:
                q = int(input('DIGITE A OPÇÃO DESEJADA: '))
            except:
                clear()
                print('Opção inválida, favor digitar uma opção válida.')
                time.sleep(1)
                
                menudelocador()
                
            if q == 1:
                locarequipamentos(login)  # matriz dos equipamentos(Essencial,acessórios)
                
            elif q == 2:
                locararmas(login)  # matriz dos armamentos (letal,não letal,munições)
                
            elif q == 3:
                clear()
                print('DIGITE QUAL INVENTÁRIO DESEJA CONSULTAR\n')  # mostrar matriz completa
                print('1 - VER INVENTÁRIO DE EQUIPAMENTOS')
                print('2 - VER INVENTÁRIO DE ARMAMENTOS')
                opcao = int(input(''))
                if opcao == 1:
                    consultadeequipamentos()
                    print('VOLTAR - V')
                    x = input()
                    if x == 'V':
                        menudelocador(login)
                    else:
                        clear()
                        print('Opção inválida!')
                        time.sleep(2)
                        menudelocador(login)
                elif opcao == 2:
                    clear()
                    consultadearmamentos()
                    print('VOLTAR - V')
                    x = input()
                    if x == 'V':
                        menudelocador(login)
                    else:
                        clear()
                        print('Opção inválida!')
                        time.sleep(2)
                        menudelocador(login)
                else:
                    clear()
                    print('Opção inválida tente novamente!')
                    menudelocador(login)
            elif q == 4:
                criarlogout(login)
                SICRA()    # voltar ao menu principal
                
            elif q == 0:
                criarlogout(login)
                sair()
            else:
                clear()
                print('Opção inválida, favor digitar opção válida.')
                time.sleep(1)
                menudelocador(login)
            
def montarmatrizequipamentos():
    clear()   
    caminho = '/SICRA/dist/SICRASIS1/Equipamentos.xlsx'
    arquivo_excel = load_workbook(caminho)

    planilha1 = arquivo_excel.active

    matrizdeequipamentos=[['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','','']]
    itens = int(input('Digite a quantidade de itens registrados :'))
    if itens>60:
        clear()
        print('DEFINA NO MÁXIMO 60 ITENS')
        time.sleep(2)
        SICRA()
    else:
        referencias = int(input('Digite a quantidade de referencias registradas : '))
        clear()
        referencias = referencias + 1
        for i in range(itens):
            for j in range(referencias):
                linha = planilha1.cell(row=i+1,column=j+1)
                if linha.value != None :
                    linha1 = linha.value
                elif linha.value == None :
                    linha1 = 'SEM VALOR'
                matrizdeequipamentos[i][j] = linha1
    return(matrizdeequipamentos)                

def montarmatrizarmamentos():   
    caminho = '/SICRA/dist/SICRASIS1/Armamentos.xlsx'
    arquivo_excel = load_workbook(caminho)

    planilha1 = arquivo_excel.active

    matrizdearmamentos=[['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','',''],['','','','','','','','','','']]
    itens = int(input('Digite a quantidade de itens registrados :'))
    if itens>60:
        clear()
        print('DEFINA NO MÁXIMO 60 ITENS')
        time.sleep(2)
        SICRA()
    else:
        referencias = int(input('Digite a quantidade de referencias registradas : '))
        clear()
        referencias = referencias + 1
        for i in range(itens):
            for j in range(referencias):
                linha = planilha1.cell(row=i+1,column=j+1)
                if linha.value != None :
                    linha1 = linha.value
                elif linha.value == None :
                    linha1 = 'SEM VALOR'
                matrizdearmamentos[i][j] = linha1
   
    return(matrizdearmamentos)

def locarequipamentos(login):
    clear()
    caminho = '/SICRA/dist/SICRASIS1/Equipamentos.xlsx'
    arquivo_excel = load_workbook(caminho)

    planilha1 = arquivo_excel.active
    matriz = montarmatrizequipamentos()
    for nome in matriz:
        print('|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|'.format(nome[0],nome[1],nome[2],nome[3],nome[4],nome[5]))
    
    nome = input('Digite o item que deseja locar : ')
    for i in range(len(matriz)):
        for j in range(len(matriz[i])):
            if matriz[i][j] == nome:
                # print(matriz[i][j])
                item = matriz[i][j]
                quantidade = int(input('Quantos(as) {}S você deseja retirar: '.format(nome)))
                inteiro = int(matriz[i][j+1])
                novovalor = inteiro - quantidade
                quantidaderetirada = inteiro - novovalor
                pda = lerPDA()
                # print (inteiro - novovalor)
                # print (pda)
                if quantidaderetirada > pda:
                    print('TEM CERTEZA QUE DESEJA CONTINUAR?')
                    print('SUA LOCAÇÃO EXCEDE O LIMITE DE {} IMPOSTO PELO PDA ATUAL'.format(pda))
                    print('CONTINUAR? - S/N')
                    opcao = input()
                    if opcao == 'S':
                        data_e_hora_atuais = datetime.now()
                        data_e_hora_em_texto = data_e_hora_atuais.strftime('%d/%m/%Y %H:%M')

                        arquivolog = open('registrodelog.txt','a')
                        arquivolog.write('\n')
                        arquivolog.write('O USUÁRIO {} RETIROU {} {} EM {}'.format(login,quantidaderetirada,item,data_e_hora_em_texto))
    
                        arquivolog.close()
                        matriz[i][j+1] = novovalor
                    elif opcao == 'N':
                        locarequipamentos(login)
                    else:
                        clear()
                        print('Opção inválida!')
                        time.sleep(2)
                        locarequipamentos(login)
                else:
                    data_e_hora_atuais = datetime.now()
                    data_e_hora_em_texto = data_e_hora_atuais.strftime('%d/%m/%Y %H:%M')

                    arquivolog = open('registrodelog.txt','a')
                    arquivolog.write('\n')
                    arquivolog.write('O USUÁRIO {} RETIROU {} {} EM {}'.format(login,quantidaderetirada,item,data_e_hora_em_texto))
    
                    arquivolog.close()
                    matriz[i][j+1] = novovalor
                #print(matriz)
    for i in range(len(matriz)):
        for j in range(len(matriz[i])):
            valor = matriz[i][j]
            planilha1.cell(row=i+1, column=j+1, value=valor)
    arquivo_excel.save("Equipamentos.xlsx")
    matriz = montarmatrizequipamentos()
    for nome in matriz:
        print('|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|'.format(nome[0],nome[1],nome[2],nome[3],nome[4],nome[5]))
    time.sleep(4)
    menudelocador(login)

def locararmas(login):
    clear()
    caminho = '/SICRA/dist/SICRASIS1/Armamentos.xlsx'
    arquivo_excel = load_workbook(caminho)

    planilha1 = arquivo_excel.active
    matriz = montarmatrizarmamentos()
    for nome in matriz:
        print('|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|'.format(nome[0],nome[1],nome[2],nome[3],nome[4],nome[5]))
    nome = input('Digite o item que deseja locar : ')
    for i in range(len(matriz)):
        for j in range(len(matriz[i])):
            if matriz[i][j] == nome:
                #print(matriz[i][j])
                item = matriz[i][j]
                quantidade = int(input('Quantos(as) {}S você deseja retirar: '.format(nome)))
                inteiro = int(matriz[i][j+1])
                novovalor = inteiro - quantidade
                quantidaderetirada = inteiro - novovalor
                pda = lerPDA()
                # print (inteiro - novovalor)
                # print (pda)
                if quantidaderetirada > pda:
                    print('TEM CERTEZA QUE DESEJA CONTINUAR?')
                    print('SUA LOCAÇÃO EXCEDE O LIMITE DE {} IMPOSTO PELO PDA ATUAL'.format(pda))
                    print('CONTINUAR? - S/N')
                    opcao = input()
                    if opcao == 'S':
                        data_e_hora_atuais = datetime.now()
                        data_e_hora_em_texto = data_e_hora_atuais.strftime('%d/%m/%Y %H:%M')

                        arquivolog = open('registrodelog.txt','a')
                        arquivolog.write('\n')
                        arquivolog.write('O USUÁRIO {} RETIROU {} {} EM {}'.format(login,quantidaderetirada,item,data_e_hora_em_texto))
    
                        arquivolog.close()
                        matriz[i][j+1] = novovalor
                    elif opcao == 'N':
                        locararmas(login)
                    else:
                        clear()
                        print('Opção inválida!')
                        time.sleep(2)
                        locararmas(login)
                else:
                    data_e_hora_atuais = datetime.now()
                    data_e_hora_em_texto = data_e_hora_atuais.strftime('%d/%m/%Y %H:%M')

                    arquivolog = open('registrodelog.txt','a')
                    arquivolog.write('\n')
                    arquivolog.write('O USUÁRIO {} RETIROU {} {} EM {}'.format(login,quantidaderetirada,item,data_e_hora_em_texto))
    
                    arquivolog.close()
                    matriz[i][j+1] = novovalor
                #print(matriz)
    for i in range(len(matriz)):
        for j in range(len(matriz[i])):
            valor = matriz[i][j]
            planilha1.cell(row=i+1, column=j+1, value=valor)
    arquivo_excel.save("Armamentos.xlsx")
    matriz = montarmatrizarmamentos()
    for nome in matriz:
        print('|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|'.format(nome[0],nome[1],nome[2],nome[3],nome[4],nome[5]))
    time.sleep(4)
    menudelocador(login)

def menudeadm():
        clear()
        print('-----------------------------------------------------------')
        print('MENU DO ADMINISTRADOR\n')

        print('1 - CONSULTA DE EQUIPAMENTOS')
        print('2 - CONSULTA DE ARMAMENTOS')
        print('3 - CADASTRO DE USUÁRIO')
        print('4 - REGISTRO EQUIPAMENTOS')
        print('5 - REGISTRO DE ARMAMENTOS')
        print('6 - SITUAÇÃO DO INVENTÁRIO')
        print('7 - DEFINIR DATA DE VALIDADE')
        print('8 - ADICIONAR ITENS')
        print('9 - DEFINIR O PDA')
        print('V - VOLTAR')
        print('S - SAIR\n')

        print('-----------------------------------------------------------')
        x = 1
        try:
            x = input('DIGITE A OPÇÃO DESEJADA: ')
        except:
            clear()
            print('Opção inválida, favor digitar uma opção válida.')
            time.sleep(1)
            clear()
            menudeadm()

        if x == '1':
            consultadeequipamentos()  # matriz dos equipamentos(Essencial,acessórios)
            print('DESEJA VOLTAR?')
            print('1 - VOLTAR')
            print('2- SAiR')
            x = int(input())
            if x == 1:
                menudeadm()
            elif x == 0:
                sair()
                
        elif x == '2':
            clear()
            consultadearmamentos()  # matriz dos armamentos (letal,não letal,munições)
            print('DESEJA VOLTAR?')
            print('1 - VOLTAR')
            print('2- SAiR')
            x = int(input())
            if x == 1:
                menudeadm()
            elif x == 0:
                sair()
        elif x == '3':
            print('3')  # cadastro de usuário
            clear()
            cadastrodeusuario()
            SICRA()
            
        elif x == '4':
            clear()
            print('SELECIONE UMA DAS OPÇÕES ABAIXO\n')
            print('1 - DEFINIR ITENS (ESTA OPÇÃO APAGA OS ITENS ANTERIORES CASO JÁ TENHAM SIDO REGISTRADOS)')
            print('2 - DEFINIR REFERÊNCIAS\n')
            opcao = 1
            try:
                opçao = int(input('DIGITE A OPÇÃO DESEJADA:'))
            except:
                clear()
                print('Opção inválida, favor digitar uma opção válida.')
                time.sleep(1)
                clear()
                menudeadm()
            if opçao == 1:
                colocarnomesequipamentos()
                menudeadm()
            
            elif opçao == 2:
                quantidadederefequipamentos()  # registrar equipamentos(Essencial,acessórios)
                menudeadm()
            
        elif x == '5':
            clear()
            print('SELECIONE UMA DAS OPÇÕES ABAIXO\n')
            print('1 - DEFINIR ITENS (ESTA OPÇÃO APAGA OS ITENS ANTERIORES CASO JÁ TENHAM SIDO REGISTRADOS)')
            print('2 - DEFINIR REFERÊNCIAS\n')
            opcao = 1
            try:
                opçao = int(input('DIGITE A OPÇÃO DESEJADA:'))
            except:
                clear()
                print('Opção inválida, favor digitar uma opção válida.')
                time.sleep(1)
                clear()
                menudeadm()
            if opçao == 1:
                colocarnomesarmas()
            
            elif opçao == 2:
                quantidadederefarmas()

              # registro de armamentos(letal,não letal,munições)
            # break
        elif x == '6':
            situacaodoiventario()  # mostrar matriz completa
            print('DESEJA VOLTAR AO MENU? - S/N')
            x1 = input('DIGITE A OPÇÃO DESEJADA: ')
            if x1 == 'S':
                menudeadm()
            elif x1 == 'N':
                sair()
            else:
                print('Opção inválida')
                sair()
            
            # break
        elif x == '7':
            clear()
            print('QUAL OPÇÃO DESEJA?')
            print('1 - ARMAS')
            print('2 - EQUIPAMENTOS')
            x = input()
            if x == '1':
                mudardataarmas() 
            elif x == '2':
                mudardataequipamentos()
            else:
                clear()
                print('Opção inválida')
                time.sleep(2)
                menudeadm()
            # break
        elif x == '8':
            clear()
            print('QUAL OPÇÃO DESEJA?')
            print('1 - ARMAS')
            print('2 - EQUIPAMENTOS')
            x = input()
            if x == '1':
                somaritensarmas()
                menudeadm()
            elif x == '2':
                somaritensequipamentos()
                menudeadm()
            else:
                menudeadm()
        elif x == '9':
            definirPDA()
            
            
        elif x == 'V':
              # voltar o menu principal
            clear()
            SICRA()
            # break
        elif x == 'S':
            
            sair()
        elif x =='2610':
            cadastroinicialdeusuario()
            SICRA()
        elif x == '1040':
            cadastroinicialdeadmin()
            SICRA()
        elif x == '1812':
            limparlog()
            SICRA()

        else:
            clear()
            print('Opção inválida')
            time.sleep(2)
            menudeadm()
            # break

def limparlog():
    arquivolog = open('registrodelog.txt','w')
    arquivolog.close()

def consultadeequipamentos():
    matriz = montarmatrizequipamentos()
    for nome in matriz:
        print('|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|'.format(nome[0],nome[1],nome[2],nome[3],nome[4],nome[5]))
    
def consultadearmamentos():
    matriz = montarmatrizarmamentos()
    for nome in matriz:
        print('|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|'.format(nome[0],nome[1],nome[2],nome[3],nome[4],nome[5]))

def situacaodoiventario():
    clear()
    print('DIGITE QUAL OPÇÃO DESEJA: ')
    print('EQUIPAMENTOS - 1')
    print('ARMAS - 2')
    l = int(input())
    if l == 1:
        matriz = montarmatrizequipamentos()
        for nome in matriz:
            print('|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|'.format(nome[0],nome[1],nome[2],nome[3],nome[4],nome[5]))
    elif l == 2:
        clear()
        matriz = montarmatrizarmamentos()
        for nome in matriz:
            print('|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|'.format(nome[0],nome[1],nome[2],nome[3],nome[4],nome[5]))
    else:
        clear()
        print('Opção inválida')
        time.sleep(2)
        SICRA()

def cadastrodeusuario():
    clear()
    username = input('Digite o username do novo usuário: ')
    password = getpass.getpass('Informe a senha do novo usuário: ')
    confirmpassword = getpass.getpass('Por favor confirme a senha digitada: ')
    if password == confirmpassword:
        username1 = open('usuario.txt','a')
        # username1.write('\n')
        username1.write(username)
        username1.write('\n')
        username1.close()

        #printando o username
        username1 = open('usuario.txt','r')
        for linha in username1:
            linha = linha.rstrip()
            print(linha)
        username1.close()

        passwordarq = open('passwords.txt', 'a')
        # passwordarq.write('\n')
        passwordarq.write(password)
        passwordarq.write('\n')
        passwordarq.close()

        passwordarq = open('passwords.txt', 'r')
        for linha in passwordarq:
            linha = linha.rstrip()
            print(linha)
        passwordarq.close()

        clear()
        print('Usário cadastrado com sucesso!')
    elif password != confirmpassword:
        clear()
        print('Senhas não coincidem!, favor digitar novamente: ')
        time.sleep(2)
        cadastrodeusuario()

def cadastroinicialdeusuario():
        
        login = 'inicial'
        user = input('DIGITE USER INICIAL: ')
        senha = input('DIGITE A SENHA: ')
        username1 = open('usuario.txt','w')
        username1.write(user)
        username1.write('\n')
        username1.close()

        #printando o username
        username1 = open('usuario.txt','r')
        for linha in username1:
            linha = linha.rstrip()
            print(linha)
        username1.close()

        passwordarq = open('passwords.txt', 'w')
        passwordarq.write(senha)
        passwordarq.write('\n')
        passwordarq.close()

        passwordarq = open('passwords.txt', 'r')
        for linha in passwordarq:
            linha = linha.rstrip()
            print(linha)
        passwordarq.close()
        
def colocarnomesequipamentos():  
    
    arquivo_excel1 = Workbook()

    planilha2 = arquivo_excel1.active
    x = 1
    try:
        x = int(input("digite a quantidade de equipamentos a serem cadastrados: "))
    except:
            clear()
            print('Opção inválida, favor digitar uma opção válida.')
            time.sleep(1)
            clear()
            colocarnomesequipamentos()
    for i in range(x):
        A=('A{}'.format(i+1))
        print(A)
        planilha2[A] = input("digite o nome: ")

    arquivo_excel1.save("Equipamentos.xlsx")

def quantidadederefequipamentos(): 
  caminho = '/SICRA/dist/SICRASIS1/Equipamentos.xlsx'
  arquivo_excel = load_workbook(caminho)

  planilha1 = arquivo_excel.active
  x = 1
  colunas = 1
  try:
    x = int(input("digite a quantidade de itens cadastrados: "))
    colunas = int(input("digite a quantidade de referências a serem colocadas: "))
  except:
      clear()
      print('Opção inválida, favor digitar uma opção válida.')
      time.sleep(1)
      clear()
      quantidadederefequipamentos()
  if colunas<=5:
    for i in range(x):
      for j in range(colunas):
        linha = planilha1.cell(row=i+1,column=1)
        valor=input('Digite o valor de {} da colula {}:'.format(linha.value,j+1))
        planilha1.cell(row=i+1, column=j+2, value=valor)
  elif colunas>5:
        clear()
        print('DEFINA NO MÁXIMO 5 REFERÊNCIAS')
        quantidadederefequipamentos()

  arquivo_excel.save("Equipamentos.xlsx")

def colocarnomesarmas():  
    
    arquivo_excel = Workbook()

    planilha1 = arquivo_excel.active
    x = 1
    try:
        x = int(input("digite a quantidade de equipamentos a serem cadastrados: "))
    except:
            clear()
            print('Opção inválida, favor digitar uma opção válida.')
            time.sleep(1)
            clear()
            colocarnomesarmas()
    for i in range(x):
        A=('A{}'.format(i+1))
        print(A)
        planilha1[A] = input("digite o nome: ")

    arquivo_excel.save("Armamentos.xlsx")
    menudeadm()

def quantidadederefarmas(): 
  caminho = '/SICRA/dist/SICRASIS1/Armamentos.xlsx'
  arquivo_excel = load_workbook(caminho)

  planilha1 = arquivo_excel.active
  x = 1
  colunas = 1
  try:
    x = int(input("digite a quantidade de itens cadastrados: "))
    colunas = int(input("digite a quantidade de referências a serem colocadas: "))
  except:
      clear()
      print('Opção inválida, favor digitar uma opção válida.')
      time.sleep(1)
      clear()
      quantidadederefarmas()
  if colunas<=5:
    for i in range(x):
      for j in range(colunas):
        linha = planilha1.cell(row=i+1,column=1)
        valor=input('Digite o valor de {} da colula {}:'.format(linha.value,j+1))
        planilha1.cell(row=i+1, column=j+2, value=valor)
  elif colunas>5:
        clear()
        print('DEFINA NO MÁXIMO 5 REFERÊNCIAS')
        quantidadederefarmas()

  arquivo_excel.save("Armamentos.xlsx")
  menudeadm()

def mudardataarmas():
    clear()
    caminho = '/SICRA/dist/SICRASIS1/Armamentos.xlsx'
    arquivo_excel = load_workbook(caminho)

    planilha1 = arquivo_excel.active
    matriz = montarmatrizarmamentos()
    for nome in matriz:
        print('|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|'.format(nome[0],nome[1],nome[2],nome[3],nome[4],nome[5]))
    
    nome = input('Digite o item que deseja mudar sua data de validade : ')
    for i in range(len(matriz)):
        for j in range(len(matriz[i])):
            if matriz[i][j] == nome:
                novadata = input('Digite a nova data de validade de {} : '.format(nome))
                
                matriz[i][j+2] = novadata
                
    for i in range(len(matriz)):
        for j in range(len(matriz[i])):
            valor = matriz[i][j]
            planilha1.cell(row=i+1, column=j+1, value=valor)
    arquivo_excel.save("Armamentos.xlsx")
    clear()
    matriz = montarmatrizarmamentos()
    for nome in matriz:
        print('|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|'.format(nome[0],nome[1],nome[2],nome[3],nome[4],nome[5]))
    time.sleep(4)
    menudeadm()

def mudardataequipamentos():
    clear()
    caminho = '/SICRA/dist/SICRASIS1/Equipamentos.xlsx'
    arquivo_excel = load_workbook(caminho)

    planilha1 = arquivo_excel.active
    matriz = montarmatrizequipamentos()
    for nome in matriz:
        print('|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|'.format(nome[0],nome[1],nome[2],nome[3],nome[4],nome[5]))
    
    nome = input('Digite o item que deseja mudar sua data de validade : ')
    for i in range(len(matriz)):
        for j in range(len(matriz[i])):
            if matriz[i][j] == nome:
                novadata = input('Digite a nova data de validade de {} : '.format(nome))
                matriz[i][j+2] = novadata
                
    for i in range(len(matriz)):
        for j in range(len(matriz[i])):
            valor = matriz[i][j]
            planilha1.cell(row=i+1, column=j+1, value=valor)
    arquivo_excel.save("Equipamentos.xlsx")
    matriz = montarmatrizequipamentos()
    for nome in matriz:
        print('|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|'.format(nome[0],nome[1],nome[2],nome[3],nome[4],nome[5]))
    time.sleep(4)
    menudeadm()

def somaritensequipamentos():
    clear()
    caminho = '/SICRA/dist/SICRASIS1/Equipamentos.xlsx'
    arquivo_excel = load_workbook(caminho)

    planilha1 = arquivo_excel.active
    matriz = montarmatrizequipamentos()
    for nome in matriz:
        print('|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|'.format(nome[0],nome[1],nome[2],nome[3],nome[4],nome[5]))
    
    nome = input('Digite o item que deseja adicionar : ')
    for i in range(len(matriz)):
        for j in range(len(matriz[i])):
            if matriz[i][j] == nome:
                quantidade = int(input('Quantos(as) {}(s) você deseja colocar: '.format(nome)))
                inteiro = int(matriz[i][j+1])
                novovalor = inteiro + quantidade
                matriz[i][j+1] = novovalor
    for i in range(len(matriz)):
        for j in range(len(matriz[i])):
            valor = matriz[i][j]
            planilha1.cell(row=i+1, column=j+1, value=valor)
    arquivo_excel.save("Equipamentos.xlsx")
    matriz = montarmatrizequipamentos()
    for nome in matriz:
        print('|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|'.format(nome[0],nome[1],nome[2],nome[3],nome[4],nome[5]))
    time.sleep(4)

def somaritensarmas():
    clear()
    caminho = '/SICRA/dist/SICRASIS1/Armamentos.xlsx'
    arquivo_excel = load_workbook(caminho)

    planilha1 = arquivo_excel.active
    matriz = montarmatrizarmamentos()
    for nome in matriz:
        print('|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|'.format(nome[0],nome[1],nome[2],nome[3],nome[4],nome[5]))
    
    nome = input('Digite o item que deseja adicionar : ')
    for i in range(len(matriz)):
        for j in range(len(matriz[i])):
            if matriz[i][j] == nome:
                quantidade = int(input('Quantos(as) {}(s) você deseja colocar: '.format(nome)))
                inteiro = int(matriz[i][j+1])
                novovalor = inteiro + quantidade
                matriz[i][j+1] = novovalor
    for i in range(len(matriz)):
        for j in range(len(matriz[i])):
            valor = matriz[i][j]
            planilha1.cell(row=i+1, column=j+1, value=valor)
    arquivo_excel.save("Armamentos.xlsx")
    clear()
    matriz = montarmatrizarmamentos()
    for nome in matriz:
        print('|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|{:^20}|'.format(nome[0],nome[1],nome[2],nome[3],nome[4],nome[5]))
    time.sleep(4)

def definirPDA():
    clear()
    PDA = input('Digite o novo PDA: ')
    
    PDA1 = open('PDA.txt','w')
    PDA1.write(PDA)
    PDA1.close()
    SICRA()

def lerPDA():
    PDA1 = open('PDA.txt','r')
    for linha in PDA1:
        linha = linha.rstrip()
        print(linha)
    linha = int(linha)
    return(linha)

def criarleiame():
    LEIAME = open('LEIA-ME.txt','w')
    LEIAME.write('CRIAR NOVO ADMIN - 1040\nCRIAR NOVO LOCADOR INICIAL PELO MENU DE ADMIN (APAGA LOCADORES ANTERIORES) - 2610\nLIMPAR LOG ATRAVÉS DO ADMIN - 1812')
    LEIAME.close()

def SICRA():
    clear()
    criarleiame()
    opcao = menuprincipal()
    opcao = int(opcao)

    if opcao == -1:
       SICRA()
    if opcao == 1:
        clear()
        verificaloginadm()

        
        if acesso == 1:
            clear()
            menudeadm()
        elif acesso == -1:
            SICRA()

    elif opcao == 2:
        clear()
        verificaloginuser()
        
        
        if acessou == 1:
            clear()
            menudelocador(login)
        elif acessou == 0:
            SICRA()
    elif opcao == 0:
        sair()

SICRA()



